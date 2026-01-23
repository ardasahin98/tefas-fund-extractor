import time
import os
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl

# ------------------ DEFAULT FUND LIST ------------------

DEFAULT_FONDS = [
    "TI1", "TIV", "DCB", "BGP", "ALE", "TKM", "IGL", "APT",
    "AFT", "TTA", "IOG", "GO1", "GO3", "GO4", "TE3", "TPC"
]

# ------------------ TKINTER INPUT POPUP ------------------

def get_fund_input():
    result = {"funds": None}

    def submit():
        text = text_box.get("1.0", tk.END).strip()
        if text:
            funds = [f.strip().upper() for f in text.replace(",", " ").split()]
            result["funds"] = funds
        else:
            result["funds"] = DEFAULT_FONDS
        root.destroy()

    root = tk.Tk()
    root.title("TEFAS Fund List")

    tk.Label(
        root,
        text="Enter fund codes (comma or space separated).\nLeave empty to use default list:",
        pady=10
    ).pack()

    text_box = tk.Text(root, height=6, width=40)
    text_box.pack(padx=10)

    tk.Button(root, text="Run", command=submit, pady=5).pack(pady=10)

    root.mainloop()
    return result["funds"]

fonds = get_fund_input()

# ------------------ EXCEL SETUP ------------------

today = datetime.now().strftime("%Y-%m-%d")
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
file_path = os.path.join(desktop, f"fund_values_{today}.xlsx")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)  # clear old data
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fund", "Price"])

# ------------------ SELENIUM SETUP ------------------

options = Options()
options.add_argument("--disable-blink-features=AutomationControlled")
# IMPORTANT: do NOT use headless for TEFAS

service = Service(ChromeDriverManager().install())

# ------------------ MAIN LOOP ------------------

for fond_name in fonds:
    price = ""

    try:
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(f"https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={fond_name}")

        for _ in range(20):
            try:
                price = driver.find_element(
                    By.XPATH,
                    "//*[@id='MainContent_PanelInfo']//ul/li[1]/span"
                ).text
                if price.strip():
                    break
            except:
                time.sleep(0.5)

        driver.quit()
    except:
        price = ""

    ws.append([fond_name, price])

# ------------------ SAVE ------------------

wb.save(file_path)

# ------------------ DONE MESSAGE ------------------

messagebox.showinfo(
    "Completed",
    f"Excel file created/updated:\n{file_path}"
)